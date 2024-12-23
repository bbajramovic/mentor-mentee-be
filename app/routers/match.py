from fastapi import APIRouter, UploadFile, File
from app.db import get_all_object, db
from app.utils.matching import generateGroup
from pydantic import BaseModel
import pandas as pd

router = APIRouter()

class NewMatchBody(BaseModel):
    matchName: str
    file: UploadFile = File(...)

class UpdateMatchNameBody(BaseModel):
    newName: str
    uid: str

@router.get("/get/{uid}", tags=["match"])
async def get_match(uid: str):
    # Query the match by uid
    match = db.child("matches").order_by_child("uid").equal_to(uid).get()
    val = match.val()
    if not val:
        return {}
    key = list(val.keys())[0]
    return val[key]


@router.put("/update/match_name", tags=["match"])
async def update_match_name(request: UpdateMatchNameBody):
    # Query the match by uid
    uid = request.uid
    newName = request.newName
    # db.child("matches").order_by_child("uid").equal_to(uid).update({"matchName": newName})
    matches = db.child("matches").get()
    for match in matches.each():
        if match.val().get("uid") == uid:
            key = match.key()
            db.child("matches").child(key).update({"matchName": newName})
            return {"message": "Match name updated!", "status": "success"}
    return {"message": "Match not found!", "status": "error"}
@router.delete("/delete/{uid}", tags=["match"])
async def delete_match(uid: str):
    # Delete the match
    db.child("matches").order_by_child("uid").equal_to(uid).remove()
    return {"message": "Match deleted!"}
    
@router.get("/list", tags=["match"])
async def get_all_matches():

    try: 
        matches = db.child("matches").get()
        result = []
        for match in matches.each():
            result.append(match.val())
        return result
    except Exception as e:
        return []


@router.post("/new", tags=["match"])
async def create_new_match(request: NewMatchBody):
    matchName = request.matchName
    file = request.file

    # Read the XLSX file
    df = pd.read_excel(file.file)
    data = df.to_dict(orient='records')

    # Get all mentors, all mentees
    mentees = get_all_object("mentees")
    mentors = get_all_object("mentors")

    # Filter all mentees who have been matched
    # Filter all mentor who have been matched with maximum mentee they need

    result = generateGroup(mentees, mentors, matchName)

    db.child("matches").push(result)
    return result
    

# TODO:
# Delete, Update, Create, Get all, Get one, Get by id, Get by uuid

from uuid import uuid4
import re

from openpyxl import load_workbook
from io import BytesIO
from uuid import uuid4


def split_field_to_array(field):
    # Split by ";" ", ", ",", "/"
    spliited = [x.strip() for x in re.split(r"[;,/]", field) if x.strip()]
    # Remove special characters
    spliited = [re.sub(r'[^\w\s]', '', x) for x in spliited]
    return spliited

def process_mentee(wb, sheetname=None):
    if sheetname:
        sheet = wb[sheetname]
    else:
        sheet = wb.active

    # Get second row as header
    header = [cell.value for cell in sheet[3]]

    # Get the data from the sheet
    data = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        entry = {}
        if(row[header.index("Little Buddy")] == None or row[header.index("Little Buddy")] == ""):
            continue
        entry["fullName"] = row[header.index("Little Buddy")] 
        entry["email"] = row[header.index("Email")] 
        entry["gender"] = row[header.index("Gender")]
        entry["homeTown"] = row[header.index("Hometown")]
        entry["phoneNumber"] = row[header.index("Phone")]
        entry["birthYear"] = row[header.index("Birthyear")]

        occupation = {}
        occupation["companyName"] =  ""
        occupation["position"] = ""
        occupation["employmentLevel"] = ""
        occupation["yearsOfExperience"] = ""
        occupation["employmentStatus"] = ""
        occupation["industry"] = ""

        entry["occupation"] = occupation

        education = {}
        education["currentSchool"] = row[header.index("School name")]
        education["currentSchoolYear"] = row[header.index("School Year")]
        education["latestGPA"] = row[header.index("GPA")]
        education["major"] = row[header.index("Major")]

        entry["education"] = education
        mentee = {}
        
        mentee_fields = row[header.index("Mentor Field")]
        mentee["industries"] = split_field_to_array(mentee_fields)
        mentee["softSkills"] = split_field_to_array(row[header.index("Skills")])  
        mentee["preferredForeignMentor"] = "no preference"
        mentee["preferredMentorGender"] = "any"

        entry["mentee"] = mentee

        bio = {}
        bio["favoriteBook"] = row[header.index("Books")]
        bio["favoriteMovie"] = row[header.index("Films")]
        bio["favoriteQuote"] = row[header.index("Quote")]
        bio["hobbies"] = split_field_to_array(row[header.index("Hobbies")])
        bio["selfIntroduction"] = row[header.index("Introduction")]

        entry["currentLocation"] = {
            "province": "",
            "district": "",
        }

        entry["uuid"] = str(uuid4())
        entry["id"] = row[header.index("SUN")]

        data.append(entry)

    # Export the data to a JSON file
    return data

def process_mentor(wb, sheetname=None):
    if sheetname:
        sheet = wb[sheetname]
    else:
        sheet = wb.active

    # Get second row as header
    header = [cell.value for cell in sheet[3]]

    # Get the data from the sheet
    data = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        entry = {}
        if(row[header.index("Full name")] == None or row[header.index("Full name")] == ""):
            continue
        entry["fullName"] = row[header.index("Full name")] 
        entry["email"] = row[header.index("Email")] 
        entry["gender"] = row[header.index("Gender")]
        entry["homeTown"] = row[header.index("Hometown")]
        entry["phoneNumber"] = row[header.index("Phone")]
        entry["birthYear"] = row[header.index("Birthyear")]

        occupation = {}
        occupation["companyName"] = row[header.index("Company")]
        occupation["position"] = row[header.index("Job Title")]
        occupation["employmentLevel"] = row[header.index("Level")]
        occupation["yearsOfExperience"] = row[header.index("Years of Experience")]
        occupation["employmentStatus"] = "full-time"
        occupation["industry"] = row[header.index("Job Field")]

        entry["occupation"] = occupation

        mentor = {}
        mentor_field = row[header.index("Mentor Field")]
        mentor["industries"] = split_field_to_array(mentor_field)
        mentor["softSkills"] = split_field_to_array(row[header.index("Skills")])  
        mentor["preferredMenteeGender"] = "any"
        mentor["preferredMenteeCollegeYear"] = "any"
        mentor["preferredMenteeMajor"] = mentor["industries"]

        entry["currentLocation"] = {
            "province": "",
            "district": "",
        }

        entry["mentor"] = mentor

        bio = {}
        bio["favoriteBook"] = ""
        bio["favoriteMovie"] = ""
        bio["favoriteQuote"] = ""
        bio["hobbies"] = []
        bio["selfIntroduction"] = ""

        entry["uuid"] = str(uuid4())
        entry["id"] = row[header.index("No.")]

        data.append(entry)

    # Export the data to a JSON file
    return data

@router.post("/match_by_file", tags=["match"])
async def match_by_file(data: UploadFile  = File(...), matchName: str = ""):
    try:
        # Read the file content asynchronously
        file_content = await data.read()
        
        # Use BytesIO to simulate a file object from the byte data
        file_stream = BytesIO(file_content)

        # Load the workbook using openpyxl
        wb = load_workbook(file_stream, data_only=True)

        mentors = process_mentor(wb, "Big Buddy ")
        mentees = process_mentee(wb, "Little Buddy")
        
        # Save to database
        db.child("mentors").set(mentors)
        db.child("mentees").set(mentees)

        # Make mentor-mentee match
        result = generateGroup(mentees, mentors, matchName)

        db.child("matches").push(result)


        return {"message": "Create match successfully!"}
    
    except Exception as e:
        return {"error": f"Error processing the file: {e}"}